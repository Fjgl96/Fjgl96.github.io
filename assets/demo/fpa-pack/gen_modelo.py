"""Genera Excel con formulas + PDF 1-pager + PNGs del FP&A Monthly Pack."""
import json, pathlib
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = pathlib.Path(r"C:\Users\FGUERR~1\AppData\Local\Temp\opencode\fpa-pack")
d = json.loads((BASE/"data.json").read_text(encoding="utf-8"))

M1 = d["real_h1"]/1e6; M2 = d["ppto_h1"]/1e6
DESV = d["real_h1"]-d["ppto_h1"]
PC = (d["real_h1"]/d["ppto_h1"]-1)*100
E1 = d["ebitda_real"]/1e6; E2 = d["ebitda_ppto"]/1e6
DESE = d["ebitda_real"]-d["ebitda_ppto"]
def S(v): return f"S/ {v/1e6:,.2f}M".replace(",","_").replace(".",",").replace("_",".")
def SS(v): return ("-" if v < 0 else "") + f"S/ {abs(v)/1e3:,.0f}k".replace(",",".")

AZUL = "0B6D9E"; AZUL_L = "DFF3FC"; GRIS = "F5F8FD"; VERDE = "0F6B46"; ROJO = "B00020"
H_FILL = PatternFill("solid", fgColor=AZUL); H_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
T_FONT = Font(name="Calibri", bold=True, size=11, color=AZUL)
N_FONT = Font(name="Calibri", size=10); B_FONT = Font(name="Calibri", bold=True, size=10)
MONEY = '#,##0;-#,##0;-'; PCT = '0.0%;0.0%;-'; thin = Side(style="thin", color="D9E2EC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def estilo_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = H_FILL; cell.font = H_FONT; cell.alignment = Alignment(horizontal="center", wrap_text=True)

def dinero(ws, row, col, value=None, formula=None, bold=False):
    c = ws.cell(row=row, column=col)
    if formula: c.value = formula
    else: c.value = value
    c.number_format = MONEY; c.font = B_FONT if bold else N_FONT; c.border = BORDER
    return c

wb = Workbook()

# ---------- LEEME ----------
ws = wb.active; ws.title = "LEEME"
ws["A1"] = d["empresa"]; ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=AZUL)
ws["A2"] = "FP&A Monthly Pack — corte Junio 2026 (H1) · Caso 100% sintético"
ws["A2"].font = T_FONT
ws["A4"] = ("AVISO: todas las cifras son inventadas para fines demostrativos. Ningún dato corresponde "
 "a un cliente real. Moneda: soles (PEN). Tasa IR 29.5%. TC ppto 3.72, real prom H1 3.81, cierre jun 3.84.")
ws["A4"].alignment = Alignment(wrap_text=True, vertical="top"); ws.row_dimensions[4].height = 30
indice = ["PARAM: supuestos (TC, tasas, calendario)",
 "PPTO_REAL: ventas mensuales ppto vs real ene-jun + acumulado",
 "PyG_H1: estado de resultados H1 con fórmulas, desvíos y márgenes",
 "BAL: balance dic-25 vs jun-26, cuadra Activo = Pasivo + Patrimonio",
 "KPIs: DSO/DIO/DPO/CCC y rentabilidad con semáforos y fórmulas",
 "BRIDGE: puente EBITDA ppto → real reconciliado con el PVM (margen std 24%)",
 "FLUJO: caja H1 indirecta que ata neta + WC + deuda = Δcaja -250k",
 "FORECAST_H2: deuda mensual jul-dic con pico en octubre (supuesto base)",
 "COMP_LY: H1-2025 vs H1-2026 like-for-like + canales YoY",
 "PUENTE_VENTAS: volumen, precio, mix y descuento (atribuye el desvio de ventas)",
 "DRILL_CLIENTES: CxC Constructor por cliente y aging, con DSO",
 "DRILL_SKUS: inventarios Acabados por SKU, valorizado y cobertura",
 "FX_FORECAST: exposición USD, sensibilidad y forecast FY con plan de acción",
 "DICC: diccionario de campos y linaje"]
for i, t in enumerate(indice, start=6):
    ws.cell(row=i, column=1, value=t).font = N_FONT
ws.column_dimensions["A"].width = 130
for r in [1,2,4]:
    ws.merge_cells(f"A{r}:A{r}")

# ---------- PARAM ----------
ws = wb.create_sheet("PARAM")
params = [("Parámetro","Valor","Fuente / nota"),
 ("TC presupuesto 2026",3.72,"Supuesto anual"),
 ("TC promedio real H1",3.81,"BCRP (sintético)"),
 ("TC cierre junio",3.84,"BCRP (sintético)"),
 ("Tasa IR",0.295,"Régimen general"),
 ("Días H1 (ene-jun)",181,"Calendario"),
 ("% compras en USD",0.70,"Política importaciones"),
 ("Caja mínima operativa",900_000,"Tesorería"),
 ("Presupuesto ventas FY",d["fy"]["ventas_ppto"],"PPTO_2026 aprobado"),
 ("Línea revolvente propuesta",d["fy"]["linea_adicional"],"Banco, por aprobar")]
for r,(a,b,c_) in enumerate([("Parámetro","Valor","Fuente / nota")]+params[0:], start=1):
    for col,v in enumerate([a,b,c_], start=1):
        cell = ws.cell(row=r, column=col, value=v); cell.border = BORDER
        cell.font = H_FONT if r==1 else N_FONT
        if r==1: cell.fill = H_FILL
estilo_header(ws,1,3)
ws["B3"].number_format='0.00'; ws["B4"].number_format='0.00'; ws["B5"].number_format='0.00'
ws["B7"].number_format=PCT; ws["B10"].number_format=MONEY; ws["B11"].number_format=MONEY
for col,w in zip("ABC",(34,22,44)): ws.column_dimensions[col].width=w

# ---------- PPTO_REAL ----------
ws = wb.create_sheet("PPTO_REAL")
hdr = ["Mes","Ppto","Real","Desvío (S/)","Desvío %"]
for c,h in enumerate(hdr,1): ws.cell(row=1,column=c,value=h)
estilo_header(ws,1,5)
for i,m in enumerate(d["meses"][:6], start=2):
    p = d["ppto_m"][i-2]; r_ = d["real_m"][i-2]
    ws.cell(row=i,column=1,value=m).font=N_FONT
    dinero(ws,i,2,p); dinero(ws,i,3,r_)
    dinero(ws,i,4,None,formula=f"=C{i}-B{i}")
    c5 = ws.cell(row=i,column=5); c5.value=f"=IF(B{i}=0,0,C{i}/B{i}-1)"; c5.number_format=PCT; c5.font=N_FONT; c5.border=BORDER
tot = 8
ws.cell(row=tot,column=1,value="H1").font=B_FONT
for col in (2,3): dinero(ws,tot,col,None,formula=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}7)",bold=True)
dinero(ws,tot,4,None,formula="=C8-B8",bold=True)
c5 = ws.cell(row=tot,column=5); c5.value="=C8/B8-1"; c5.number_format=PCT; c5.font=B_FONT; c5.border=BORDER
for col,w in zip("ABCDE",(10,16,16,16,12)): ws.column_dimensions[col].width=w

# ---------- PyG_H1 ----------
ws = wb.create_sheet("PyG_H1")
for c,h in enumerate(["Cuenta","Ppto H1","Real H1","Desvío","Desv %","Margen real"],1):
    ws.cell(row=1,column=c,value=h)
estilo_header(ws,1,6)
filas = [("Ventas netas",d["ppto_h1"],d["real_h1"],False),
 ("Costo de ventas",d["cogs_ppto"],d["cogs_real"],False),
 ("Utilidad bruta",d["ub_ppto"],d["ub_real"],True),
 ("Gastos operativos",d["opex_ppto"],d["opex_real"],False),
 ("EBITDA",d["ebitda_ppto"],d["ebitda_real"],True),
 ("Depreciación",DEP:=450_000,450_000,False),
 ("EBIT",d["ebit_ppto"],d["ebit_real"],True),
 ("Gastos financieros",d["gf_ppto"],d["gf_real"],False),
 ("Diferencia de cambio",d["dc_ppto"],d["dc_real"],False),
 ("EBT",d["ebt_ppto"],d["ebt_real"],True),
 ("Impuesto a la renta 29.5%",d["ir_ppto"],d["ir_real"],False),
 ("Utilidad neta",d["net_ppto"],d["net_real"],True)]
for i,(cta,p,r_,bold) in enumerate(filas, start=2):
    ws.cell(row=i,column=1,value=cta).font = B_FONT if bold else N_FONT
    dinero(ws,i,2,p,bold=bold); dinero(ws,i,3,r_,bold=bold)
    dinero(ws,i,4,None,formula=f"=C{i}-B{i}",bold=bold)
    c5=ws.cell(row=i,column=5); c5.value=f"=IF(B{i}=0,0,C{i}/B{i}-1)"; c5.number_format=PCT; c5.font=B_FONT if bold else N_FONT; c5.border=BORDER
    c6=ws.cell(row=i,column=6)
    if cta in ("Utilidad bruta","EBITDA","EBIT","EBT","Utilidad neta"):
        c6.value=f"=IF($C$2=0,0,C{i}/$C$2)"; c6.number_format=PCT
    else: c6.value="—"
    c6.font=B_FONT if bold else N_FONT; c6.border=BORDER
for col,w in zip("ABCDEF",(30,16,16,16,10,12)): ws.column_dimensions[col].width=w

# ---------- BAL ----------
ws = wb.create_sheet("BAL")
for c,h in enumerate(["Cuenta","Dic-25","Jun-26","Var S/","Var %"],1): ws.cell(row=1,column=c,value=h)
estilo_header(ws,1,5)
orden = [("Caja",1),("Cuentas por cobrar",1),("Inventarios",1),("Otros AC",1),("ACTIVO CORRIENTE",2),
 ("Activo fijo neto",1),("TOTAL ACTIVO",2),("Cuentas por pagar comerciales",1),
 ("Deuda CP",1),("Otros pasivos CP",1),("TOTAL PASIVO CORRIENTE",2),("Deuda LP",1),
 ("PATRIMONIO",1),("TOTAL PAS+PAT",2)]
Dic = d["bal_dic"]; Jun = d["bal_jun"]
vals = {"Caja":(Dic["caja"],Jun["caja"]),"Cuentas por cobrar":(Dic["cxc"],Jun["cxc"]),
 "Inventarios":(Dic["inv"],Jun["inv"]),"Otros AC":(Dic["otros_ac"],Jun["otros_ac"]),
 "ACTIVO CORRIENTE":(Dic["caja"]+Dic["cxc"]+Dic["inv"]+Dic["otros_ac"],Jun["caja"]+Jun["cxc"]+Jun["inv"]+Jun["otros_ac"]),
 "Activo fijo neto":(Dic["af_neto"],Jun["af_neto"]),
 "TOTAL ACTIVO":(Dic["caja"]+Dic["cxc"]+Dic["inv"]+Dic["otros_ac"]+Dic["af_neto"],Jun["caja"]+Jun["cxc"]+Jun["inv"]+Jun["otros_ac"]+Jun["af_neto"]),
 "Cuentas por pagar comerciales":(Dic["cxp"],Jun["cxp"]),"Deuda CP":(Dic["deuda_cp"],Jun["deuda_cp"]),
 "Otros pasivos CP":(Dic["otros_pc"],Jun["otros_pc"]),
 "TOTAL PASIVO CORRIENTE":(Dic["cxp"]+Dic["deuda_cp"]+Dic["otros_pc"],Jun["cxp"]+Jun["deuda_cp"]+Jun["otros_pc"]),
 "Deuda LP":(Dic["deuda_lp"],Jun["deuda_lp"]),"PATRIMONIO":(Dic["patrimonio"],Jun["patrimonio"]),
 "TOTAL PAS+PAT":(Dic["cxp"]+Dic["deuda_cp"]+Dic["otros_pc"]+Dic["deuda_lp"]+Dic["patrimonio"],
                  Jun["cxp"]+Jun["deuda_cp"]+Jun["deuda_cp"]*0+Jun["otros_pc"]+Jun["deuda_lp"]+Jun["patrimonio"])}
r = 2
for cta,_ in orden:
    a,b = vals[cta]; bold = cta.isupper()
    ws.cell(row=r,column=1,value=cta).font = B_FONT if bold else N_FONT
    dinero(ws,r,2,a,bold=bold); dinero(ws,r,3,b,bold=bold)
    dinero(ws,r,4,None,formula=f"=C{r}-B{r}",bold=bold)
    c5=ws.cell(row=r,column=5); c5.value=f"=IF(B{r}=0,0,C{r}/B{r}-1)"; c5.number_format=PCT; c5.font=B_FONT if bold else N_FONT; c5.border=BORDER
    r += 1
ws.cell(row=r+1,column=1,value="Check: Activo − (Pas+Pat) = 0 →").font=B_FONT
dinero(ws,r+1,2,None,formula=f"=C8-C15",bold=True)
for col,w in zip("ABCDE",(32,16,16,16,10)): ws.column_dimensions[col].width=w

# ---------- KPIs ----------
ws = wb.create_sheet("KPIs")
for c,h in enumerate(["KPI","Dic-25","Ppto H1","Real H1","Desv vs ppto","Semáforo"],1): ws.cell(row=1,column=c,value=h)
estilo_header(ws,1,6)
k = d["k_real"]; kd = d["k_dic"]; su = d["supuestos"]
krows = [("DSO (días)",kd["dso"],su["dso_ppto"],k["dso"],"alto malo"),
 ("DIO (días)",kd["dio"],58.0,k["dio"],"alto malo"),
 ("DPO (días)",kd["dpo"],34.0,k["dpo"],"alto bueno"),
 ("CCC (días)",kd["ccc"],69.0,k["ccc"],"alto malo"),
 ("Margen bruto %",None,round(d["ub_ppto"]/d["ppto_h1"]*100,1),round(d["ub_real"]/d["real_h1"]*100,1),"alto bueno"),
 ("Margen EBITDA %",None,round(d["ebitda_ppto"]/d["ppto_h1"]*100,1),round(d["ebitda_real"]/d["real_h1"]*100,1),"alto bueno"),
 ("Deuda total / EBITDA LTM (x)",1.55,1.45,su["deuda_ebitda"],"alto malo"),
 ("Cobertura intereses (x)",4.6,4.2,su["cobertura"],"alto bueno")]
for i,(n,dic_,pp,real,crit) in enumerate(krows, start=2):
    ws.cell(row=i,column=1,value=n).font=N_FONT
    ws.cell(row=i,column=2,value=dic_).font=N_FONT
    ws.cell(row=i,column=3,value=pp).font=N_FONT
    ws.cell(row=i,column=4,value=real).font=B_FONT
    c5=ws.cell(row=i,column=5); c5.value=f"=D{i}-C{i}"; c5.number_format='0.0'; c5.font=N_FONT; c5.border=BORDER
    for col in (2,3,4): ws.cell(row=i,column=col).border=BORDER
    sem = ws.cell(row=i,column=6)
    if "bueno" in crit: sem.value = "OK" if real>=pp or (crit=="alto malo" and real<=pp) else "Alerta"
    else: sem.value = "Alerta" if (crit=="alto malo" and real>pp) else "OK"
    sem.font=B_FONT; sem.border=BORDER
    # correccion simple: para alto-bueno OK si real>=pp; alto-malo OK si real<=pp
    if crit=="alto bueno": sem.value = "OK" if real>=pp else "Alerta"
    else: sem.value = "OK" if real<=pp else "Alerta"
for col,w in zip("ABCDEF",(30,12,12,12,14,12)): ws.column_dimensions[col].width=w

# ---------- BRIDGE ----------
ws = wb.create_sheet("BRIDGE")
for c,h in enumerate(["Paso","Monto (S/)","Acumulado"],1): ws.cell(row=1,column=c,value=h)
estilo_header(ws,1,3)
acum=0; n=len(d["bridge"])
for i,b in enumerate(d["bridge"], start=2):
    last=(i==n+1); first=(i==2)
    ws.cell(row=i,column=1,value=b["paso"]).font = B_FONT if (first or last) else N_FONT
    dinero(ws,i,2,b["monto"],bold=(first or last))
    if first: acum=b["monto"]
    elif not last: acum+=b["monto"]
    else: acum=b["monto"]
    dinero(ws,i,3,acum,bold=(first or last))
for col,w in zip("ABC",(40,16,16)): ws.column_dimensions[col].width=w

# ---------- FLUJO (V4) ----------
ws = wb.create_sheet("FLUJO")
ws["A1"]="Flujo de caja H1 indirecto (S/) — ata Δcaja -250k"; ws["A1"].font=T_FONT
for c,h in enumerate(["Concepto","Monto","Check"],1): ws.cell(row=2,column=c,value=h)
estilo_header(ws,2,3)
fl = d["flujo"]
frowns = [("Utilidad neta",fl["neta"]),("Depreciación",fl["dep"]),
 ("Δ Cuentas por cobrar",fl["d_cxc"]),("Δ Inventarios",fl["d_inv"]),
 ("Δ Cuentas por pagar comerciales",fl["d_cxp_com"]),("Δ Otros pasivos CP",fl["d_otros_pc"]),
 ("FLUJO OPERATIVO",fl["fco"]),("Capex",fl["capex"]),
 ("Δ Deuda CP",fl["d_deuda_cp"]),("Δ Deuda LP",fl["d_deuda_lp"]),
 ("Δ CAJA (dic->jun)",fl["d_caja"])]
for i,(a,b_) in enumerate(frowns, start=3):
    bold = a.isupper()
    ws.cell(row=i,column=1,value=a).font = B_FONT if bold else N_FONT
    dinero(ws,i,2,b_,bold=bold)
    cc=ws.cell(row=i,column=3); cc.border=BORDER; cc.font=N_FONT
    if a=="FLUJO OPERATIVO": cc.value="=SUM(B3:B8)"
    elif a=="Δ CAJA (dic->jun)": cc.value="=B9+B10+B11+B12+B13"
dinero(ws,14,2,None,formula="=C9-B9",bold=True)
ws.cell(row=14,column=1,value="Check FCO (debe dar 0)").font=B_FONT
dinero(ws,15,2,None,formula="=C13-B13",bold=True)
ws.cell(row=15,column=1,value="Check caja (debe dar 0)").font=B_FONT
for col,w in zip("ABC",(36,16,16)): ws.column_dimensions[col].width=w

# ---------- FORECAST_H2 (V4) ----------
ws = wb.create_sheet("FORECAST_H2")
ws["A1"]="Deuda total fin de mes jul-dic 2026 base (S/) — supuesto, pico octubre"; ws["A1"].font=T_FONT
for c,h in enumerate(["Mes","Deuda fin de mes"],1): ws.cell(row=2,column=c,value=h)
estilo_header(ws,2,2)
for i,(m,v) in enumerate(zip(d["meses"][6:],d["deuda_h2"]), start=3):
    ws.cell(row=i,column=1,value=m).font=N_FONT
    dinero(ws,i,2,v,bold=(v==max(d["deuda_h2"])))
for col,w in zip("AB",(12,20)): ws.column_dimensions[col].width=w

# ---------- FX_FORECAST ----------
ws = wb.create_sheet("FX_FORECAST")
ws["A1"]="Exposición cambiaria (jun-26)"; ws["A1"].font=T_FONT
fxrows = [("CxP exterior (USD)",d["fx"]["cxp_usd"]),("Caja USD",-d["fx"]["caja_usd"]),
 ("Exposición neta pasiva (USD)",d["fx"]["exposicion_usd"]),
 ("Sensibilidad por 10 cts (PEN)",d["fx"]["sens_10cts"]),
 ("% sobre utilidad neta H1",d["fx"]["pct_neta"])]
for i,(a,b) in enumerate(fxrows, start=2):
    ws.cell(row=i,column=1,value=a).font=N_FONT
    dinero(ws,i,2,b)
ws["A8"]="Forecast FY26 — escenario base con plan de acción"; ws["A8"].font=T_FONT
for c,h in enumerate(["Concepto","Ppto FY","Base FY","Desvío"],1): ws.cell(row=9,column=c,value=h)
estilo_header(ws,9,4)
frowns = [("Ventas",d["fy"]["ventas_ppto"],d["fy"]["ventas_base"]),
 ("EBITDA",d["fy"]["ebitda_ppto"],d["fy"]["ebitda_base"]),
 ("Utilidad neta",d["fy"]["neta_ppto"],d["fy"]["neta_base"])]
for i,(a,p,b_) in enumerate(frowns, start=10):
    ws.cell(row=i,column=1,value=a).font=N_FONT
    dinero(ws,i,2,p); dinero(ws,i,3,b_); dinero(ws,i,4,None,formula=f"=C{i}-B{i}")
ws.cell(row=13,column=1,value="Pico deuda octubre (base)").font=N_FONT; dinero(ws,13,2,d["fy"]["pico_deuda_oct"])
ws.cell(row=14,column=1,value="Línea adicional propuesta").font=N_FONT; dinero(ws,14,2,d["fy"]["linea_adicional"])
ws.cell(row=16,column=1,value="Plan de acción (efectos 2H)").font=T_FONT
for c,h in enumerate(["ID","Acción","Efecto caja","Efecto PyG","KPI"],1): ws.cell(row=17,column=c,value=h)
estilo_header(ws,17,5)
for i,a in enumerate(d["acciones"], start=18):
    ws.cell(row=i,column=1,value=a["id"]).font=N_FONT
    ws.cell(row=i,column=2,value=a["nombre"]).font=N_FONT
    dinero(ws,i,3,a["efecto_caja"])
    ws.cell(row=i,column=4,value=a["efecto_py"]).font=N_FONT
    ws.cell(row=i,column=5,value=a["kpi"]).font=N_FONT
for col,w in zip("ABCDE",(34,42,16,30,30)): ws.column_dimensions[col].width=w

# ---------- COMP_LY (V2) ----------
ws = wb.create_sheet("COMP_LY")
ws["A1"]="Comparativo H1-2025 vs H1-2026 (S/) — like-for-like"; ws["A1"].font=T_FONT
for c,h in enumerate(["Cuenta","H1-2025","H1-2026 real","Var YoY","Var %"],1): ws.cell(row=2,column=c,value=h)
estilo_header(ws,2,5)
lyrows = [("Ventas netas",d["ly_h1"],d["real_h1"],0),("Costo de ventas",d["ly_cogs"],d["cogs_real"],0),
 ("Utilidad bruta",d["ly_ub"],d["ub_real"],1),("Gastos operativos",d["ly_opex"],d["opex_real"],0),
 ("EBITDA",d["ly_ebitda"],d["ebitda_real"],1),("Utilidad neta",d["ly_net"],d["net_real"],1)]
for i,(cta,ly,rl,bold) in enumerate(lyrows, start=3):
    ws.cell(row=i,column=1,value=cta).font = B_FONT if bold else N_FONT
    dinero(ws,i,2,ly,bold=bold); dinero(ws,i,3,rl,bold=bold)
    dinero(ws,i,4,None,formula=f"=C{i}-B{i}",bold=bold)
    c5=ws.cell(row=i,column=5); c5.value=f"=IF(B{i}=0,0,C{i}/B{i}-1)"; c5.number_format=PCT; c5.font=B_FONT if bold else N_FONT; c5.border=BORDER
r = 10
ws.cell(row=r,column=1,value="Por canal — ventas (S/)").font=T_FONT; r+=1
for c,h in enumerate(["Canal","H1-2025","H1-2026 real","Var YoY","Ppto H1","Vs ppto"],1): ws.cell(row=r,column=c,value=h)
estilo_header(ws,r,6); r+=1
for canal in list(d["canal_ppto"].keys()):
    ly=d["ly_canal"][canal]; rl=d["canal_real"][canal]; pp=d["canal_ppto"][canal]
    ws.cell(row=r,column=1,value=canal).font=N_FONT
    dinero(ws,r,2,ly); dinero(ws,r,3,rl)
    dinero(ws,r,4,None,formula=f"=C{r}-B{r}")
    dinero(ws,r,5,pp); dinero(ws,r,6,None,formula=f"=C{r}-E{r}")
    r+=1
for col,w in zip("ABCDEF",(28,16,16,16,16,16)): ws.column_dimensions[col].width=w

# ---------- PUENTE_VENTAS (V2) ----------
ws = wb.create_sheet("PUENTE_VENTAS")
for c,h in enumerate(["Paso","Monto (S/)","Acumulado"],1): ws.cell(row=1,column=c,value=h)
estilo_header(ws,1,3)
acum=0; n=len(d["pvm"])
for i,b in enumerate(d["pvm"], start=2):
    last=(i==n+1); first=(i==2)
    ws.cell(row=i,column=1,value=b["paso"]).font = B_FONT if (first or last) else N_FONT
    dinero(ws,i,2,b["monto"],bold=(first or last))
    if first: acum=b["monto"]
    elif not last: acum+=b["monto"]
    else: acum=b["monto"]
    dinero(ws,i,3,acum,bold=(first or last))
for col,w in zip("ABC",(40,16,16)): ws.column_dimensions[col].width=w

# ---------- DRILL_CLIENTES (V2) ----------
ws = wb.create_sheet("DRILL_CLIENTES")
ws["A1"]=f"Drill-down cobranza — CxC Constructor {S(d['cxc_canal']['Constructor'])} por cliente y aging (S/)"; ws["A1"].font=T_FONT
for c,h in enumerate(["Cliente","Corriente","31-60","61-90","90+","Total","DSO (d)","Mora 60+"],1):
    ws.cell(row=2,column=c,value=h)
estilo_header(ws,2,8)
for i,c in enumerate(d["clientes_inst"], start=3):
    ws.cell(row=i,column=1,value=c["nombre"]).font=N_FONT
    for col,k in [(2,"corriente"),(3,"b30"),(4,"b60"),(5,"b90")]: dinero(ws,i,col,c[k])
    dinero(ws,i,6,None,formula=f"=SUM(B{i}:E{i})",bold=True)
    cc=ws.cell(row=i,column=7,value=c["dso"]); cc.font=N_FONT; cc.border=BORDER; cc.number_format='0.0'
    dinero(ws,i,8,None,formula=f"=D{i}+E{i}")
tot=3+len(d["clientes_inst"])
ws.cell(row=tot,column=1,value="TOTAL").font=B_FONT
for col in (2,3,4,5,6,8):
    dinero(ws,tot,col,None,formula=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{tot-1})",bold=True)
for col,w in zip("ABCDEFGH",(30,14,14,14,14,14,10,14)): ws.column_dimensions[col].width=w

# ---------- DRILL_SKUS (V2) ----------
ws = wb.create_sheet("DRILL_SKUS")
ws["A1"]=f"Drill-down inventarios — linea Acabados {S(d['inv_linea']['Acabados'])} por SKU"; ws["A1"].font=T_FONT
for c,h in enumerate(["SKU","Descripcion","Stock (u)","Costo u","Valorizado","Vta/mes (u)","Cobertura (m)"],1):
    ws.cell(row=2,column=c,value=h)
estilo_header(ws,2,7)
for i,s in enumerate(d["skus_hogar"], start=3):
    ws.cell(row=i,column=1,value=s["sku"]).font=N_FONT
    ws.cell(row=i,column=2,value=s["nombre"]).font=N_FONT
    for col,k in [(3,"stock_u"),(4,"costo_u")]:
        cc=ws.cell(row=i,column=col,value=s[k]); cc.font=N_FONT; cc.border=BORDER
    dinero(ws,i,5,s["valorizado"],bold=(s["sku"]=="HOG-RESTO"))
    cc=ws.cell(row=i,column=6,value=s["vta_mes_u"]); cc.font=N_FONT; cc.border=BORDER
    cc=ws.cell(row=i,column=7,value=s["cobertura_m"]); cc.font=N_FONT; cc.border=BORDER; cc.number_format='0.0'
tot=3+len(d["skus_hogar"])
ws.cell(row=tot,column=1,value="TOTAL").font=B_FONT
dinero(ws,tot,5,None,formula=f"=SUM(E3:E{tot-1})",bold=True)
for col,w in zip("ABCDEFG",(12,30,12,10,14,12,13)): ws.column_dimensions[col].width=w

# ---------- DICC ----------
ws = wb.create_sheet("DICC")
for c,h in enumerate(["Campo","Definición","Origen"],1): ws.cell(row=1,column=c,value=h)
estilo_header(ws,1,3)
dicc = [("Ventas netas","Ventas brutas menos devoluciones y descuentos","ERP ventas (sintético)"),
 ("DSO","CxC / ventas diarias","Balance + PyG"),
 ("DIO","Inventarios / costo diario","Balance + PyG"),
 ("DPO","CxP / compras diarias","Balance + compras"),
 ("CCC","DSO + DIO − DPO","Calculado"),
 ("Exposición USD","CxP USD − caja USD","Tesorería"),
 ("Bridge EBITDA","Atribución volumen/precio/FX/descuento/opex","FP&A"),
  ("Forecast base",f"Ventas {S(d['fy']['ventas_base'])}, EBITDA {S(d['fy']['ebitda_base'])}, pico deuda oct {S(d['fy']['pico_deuda_oct'])}","Modelo + acciones A1-A5"),
  ("Flujo H1","FCO -1.13M, capex -350k, deuda neta +1.23M, caja -250k","PyG + Balance"),
  ("Deuda H2","jul 10.4 a oct 11.2 y dic 10.0 (supuesto base)","Escenario, no ERP"),
  ("H1-2025","Ventas 26.9M, EBITDA 2.76M: crece +5.6% YoY pero el margen se comprime","Comparativo like-for-like"),
  ("Puente PVM","Volumen -1.10M, precio +250k, mix -450k, descuento -400k","FP&A (ventas ppto->real)"),
 ("Aging constructor",f"{S(d['cxc_canal']['Constructor'])} en 8 clientes, mora 60+ de {SS(d['mora60'])} (DSO {d['k_real']['dso']:.0f}d del canal)".replace(".",","),"Auxiliar cobranza (sintetico)"),
 ("SKUs Acabados",f"{S(d['inv_linea']['Acabados'])} en 8 SKUs, 3 con +6 meses de cobertura ({SS(d['sobrestock'])})","Auxiliar inventarios (sintetico)")]

# LEEME: 9 -> 13 hojas
for i,(a,b,c_) in enumerate(dicc, start=2):
    for col,v in enumerate([a,b,c_],1):
        cell=ws.cell(row=i,column=col,value=v); cell.font=N_FONT; cell.border=BORDER; cell.alignment=Alignment(wrap_text=True,vertical="top")
for col,w in zip("ABC",(24,60,32)): ws.column_dimensions[col].width=w

for ws in wb.worksheets:
    ws.sheet_properties.pageSetUpPr.fitToPage = True
wb.save(BASE/"fpa_pack_jun2026.xlsx")
print("xlsx OK")

# ---------- PNGs ----------

# ---------- PNGs ----------
plt.rcParams.update({"font.size":10,"axes.spines.top":False,"axes.spines.right":False})
# 1. Ventas ppto vs real
fig,ax = plt.subplots(figsize=(9,4))
x = range(6); w=0.35
ax.bar([i-w/2 for i in x],[v/1e6 for v in d["ppto_m"][:6]],width=w,label="Ppto",color="#0B6D9E")
ax.bar([i+w/2 for i in x],[v/1e6 for v in d["real_m"][:6]],width=w,label="Real",color="#149DDD")
ax.set_xticks(list(x),d["meses"][:6]); ax.set_ylabel("S/ millones"); ax.set_title(f"Ventas H1 2026: real vs ppto ({PC:+.1f}%, {SS(DESV)})".replace(".",","))
ax.legend(frameon=False)
fig.tight_layout(); fig.savefig(BASE/"ventas_ppto_real.png",dpi=150); plt.close(fig)
# 2. Bridge waterfall simplificado
labels=[b["paso"] for b in d["bridge"]]; vals=[b["monto"]/1e3 for b in d["bridge"]]
fig,ax=plt.subplots(figsize=(10,4.2))
cum=vals[0]; xs=list(range(len(labels))); colors=[]
heights=[]; bottoms=[]
for i,v in enumerate(vals):
    if i==0 or i==len(vals)-1: heights.append(v if i==0 else v); bottoms.append(0); colors.append("#0B6D9E")
    else: bottoms.append(cum); heights.append(v); colors.append("#149DDD" if v>0 else "#C0392B"); cum+=v
ax.bar(xs,heights,bottom=bottoms,color=colors)
ax.set_xticks(xs,labels,rotation=18,ha="right",fontsize=8)
ax.set_ylabel("S/ miles"); ax.set_title(f"Puente EBITDA H1: ppto {S(d['ebitda_ppto'])} → real {S(d['ebitda_real'])} ({SS(DESE)})")
fig.tight_layout(); fig.savefig(BASE/"bridge_ebitda.png",dpi=150); plt.close(fig)
# 3. CCC
fig,ax=plt.subplots(figsize=(7,3.6))
ax.bar(["Dic-25","Real jun-26"],[d["k_dic"]["ccc"],d["k_real"]["ccc"]],color=["#90A4AE","#0B6D9E"])
ax.set_ylabel("Días"); ax.set_title(f"Ciclo de conversión: {d['k_dic']['ccc']:.0f} → {d['k_real']['ccc']:.0f} días (+{d['k_real']['ccc']-d['k_dic']['ccc']:.0f})")
for i,v in enumerate([d["k_dic"]["ccc"],d["k_real"]["ccc"]]): ax.text(i,v+1,f"{v:.0f} d",ha="center",fontweight="bold")
fig.tight_layout(); fig.savefig(BASE/"ccc.png",dpi=150); plt.close(fig)
# 4. Puente de ventas PVM (V2)
labels=[b["paso"] for b in d["pvm"]]; vals=[b["monto"]/1e3 for b in d["pvm"]]
fig,ax=plt.subplots(figsize=(10,4.2))
cum=vals[0]; heights=[]; bottoms=[]; colors=[]
for i,v in enumerate(vals):
    if i==0 or i==len(vals)-1: heights.append(v); bottoms.append(0); colors.append("#0B6D9E")
    else: bottoms.append(cum); heights.append(v); colors.append("#149DDD" if v>0 else "#C0392B"); cum+=v
ax.bar(range(len(labels)),heights,bottom=bottoms,color=colors)
ax.set_ylim(27800,30400)  # eje recortado: los deltas de ±400k son invisibles a escala completa
ax.set_xticks(range(len(labels)),[l.replace(" (+2% lista parcial)","") for l in labels],rotation=16,ha="right",fontsize=8)
ax.set_ylabel("S/ miles (eje recortado)"); ax.set_title(f"Puente de ventas H1: ppto {S(d['ppto_h1'])} → real {S(d['real_h1'])} ({SS(DESV)})")
fig.tight_layout(); fig.savefig(BASE/"pvm_ventas.png",dpi=150); plt.close(fig)
# 5. Aging institucional (V2)
cl=d["clientes_inst"]
fig,ax=plt.subplots(figsize=(10,4.4))
ypos=range(len(cl))
ax.barh(list(ypos),[c["corriente"]/1e3 for c in cl],color="#0B6D9E",label="Corriente")
ax.barh(list(ypos),[c["b30"]/1e3 for c in cl],left=[c["corriente"]/1e3 for c in cl],color="#149DDD",label="31-60")
ax.barh(list(ypos),[c["b60"]/1e3 for c in cl],left=[(c["corriente"]+c["b30"])/1e3 for c in cl],color="#E67E22",label="61-90")
ax.barh(list(ypos),[c["b90"]/1e3 for c in cl],left=[(c["corriente"]+c["b30"]+c["b60"])/1e3 for c in cl],color="#C0392B",label="90+")
ax.set_yticks(list(ypos),[c["nombre"] for c in cl],fontsize=8)
ax.set_xlabel("S/ miles"); ax.set_title(f"CxC Constructor {S(sum(d['cxc_canal'].values()) and d['cxc_canal']['Constructor'])} por cliente y aging — mora 60+: {SS(d['mora60'])}")
ax.legend(frameon=False,fontsize=8,ncol=4)
fig.tight_layout(); fig.savefig(BASE/"aging_inst.png",dpi=150); plt.close(fig)
# 6. Flujo de caja (V4): neta -> FCO -> caja
fl = d["flujo"]
fig,ax=plt.subplots(figsize=(10,4.2))
fl_labels=["Neta H1","+ Deprec.","- ΔCxC","- ΔInv","- ΔCxP","+ Otros PC","FCO","Capex","+ Deuda CP","- Deuda LP","Δ Caja"]
fl_vals=[fl["neta"]/1e3,fl["dep"]/1e3,fl["d_cxc"]/1e3,fl["d_inv"]/1e3,fl["d_cxp_com"]/1e3,
 fl["d_otros_pc"]/1e3,fl["fco"]/1e3,fl["capex"]/1e3,fl["d_deuda_cp"]/1e3,fl["d_deuda_lp"]/1e3,fl["d_caja"]/1e3]
cum=0; heights=[]; bottoms=[]; colors=[]
anchor={6,10}
for i,v in enumerate(fl_vals):
    if i in anchor: heights.append(v if i==6 else v); bottoms.append(0); colors.append("#0B6D9E" if v<0 else "#0F6B46")
    else: bottoms.append(cum); heights.append(v); colors.append("#149DDD" if v>0 else "#C0392B"); cum+=v
ax.bar(range(len(fl_labels)),heights,bottom=bottoms,color=colors)
ax.set_xticks(range(len(fl_labels)),fl_labels,rotation=20,ha="right",fontsize=8)
ax.set_ylabel("S/ miles"); ax.set_title(f"Flujo H1: neta {SS(fl['neta'])} + WC {SS(fl['d_cxc']+fl['d_inv']+fl['d_cxp_com']+fl['d_otros_pc'])} + deuda {SS(fl['d_deuda_cp']+fl['d_deuda_lp'])} = caja {SS(fl['d_caja'])}")
fig.tight_layout(); fig.savefig(BASE/"flujo_h1.png",dpi=150); plt.close(fig)
print("pngs OK")
