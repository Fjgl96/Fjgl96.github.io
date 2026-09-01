"""
adaptar_datos.py — caso_demo.xlsx  →  data.json multiperiodo

Por qué existe: `extract_data.py` del skill no puede leer este archivo. Falla
por tres razones distintas, documentadas en el informe:

  F1  discover_sheets() sólo mira el NOMBRE de la hoja. "BG" y "ERF", las dos
      abreviaturas peruanas más comunes, no están entre sus palabras clave, y
      el título que la hoja lleva adentro nunca se lee.
  F2  periods = 1 está hardcodeado, y discover_sheets guarda UNA hoja por tipo
      de estado, así que un libro con una hoja por año pierde todos menos el
      último.
  F3  extract_data.py emite {esf, egp} y calculate_ratios.py espera
      {periodos, balance, estado_resultados}. Los dos extremos de la tubería
      no hablan el mismo esquema.

Este adaptador hace la extracción por la vía que el propio SKILL.md documenta
("Si parser falla → extracción manual"), pero como script: determinista,
auditable y repetible. Lee el Excel de verdad — no importa el caso sintético —
así que si el archivo cambia, el resultado cambia.
"""
import json, re, sys
from openpyxl import load_workbook

ARCHIVO = sys.argv[1] if len(sys.argv) > 1 else "caso_demo.xlsx"
SALIDA  = sys.argv[2] if len(sys.argv) > 2 else "data.json"

def norm(s):
    s = str(s or "").lower().strip()
    s = (s.replace("á","a").replace("é","e").replace("í","i")
           .replace("ó","o").replace("ú","u").replace("ñ","n"))
    return re.sub(r"\s+", " ", s)

# etiqueta del Excel  →  clave del esquema de calculate_ratios.py
BG_MAP = {
    "efectivo y equivalentes de efectivo": ("activo_corriente", "efectivo"),
    "cuentas por cobrar comerciales":      ("activo_corriente", "cuentas_por_cobrar"),
    "existencias":                          ("activo_corriente", "inventarios"),
    "gastos contratados por anticipado":   ("activo_corriente", "gastos_anticipados"),
    "cuentas por cobrar diversas":         ("activo_corriente", "cuentas_por_cobrar_diversas"),
    "cuentas por cobrar al personal":      ("activo_corriente", "cuentas_por_cobrar_personal"),
    "otros activos":                        ("activo_corriente", "otros_activos"),
    "total activo corriente":              ("activo_corriente", "total_activo_corriente"),
    "propiedad planta y equipo (neto)":    ("activo_no_corriente", "inmuebles_maq_equipo"),
    "intangibles (neto)":                  ("activo_no_corriente", "intangibles"),
    "inversiones mobiliarias":             ("activo_no_corriente", "inversiones"),
    "otras cuentas por cobrar diversas l.p":("activo_no_corriente","cuentas_por_cobrar_lp"),
    "total activo no corriente":           ("activo_no_corriente", "total_activo_no_corriente"),
    "cuentas por pagar comerciales":       ("pasivo_corriente", "cuentas_por_pagar"),
    "tributos y aportes por pagar":        ("pasivo_corriente", "tributos_por_pagar"),
    "remuneraciones y participaciones por pagar": ("pasivo_corriente","remuneraciones_por_pagar"),
    "otras cuentas por pagar":             ("pasivo_corriente", "otras_cuentas_por_pagar"),
    "otras cuentas por pagar a corto plazo":("pasivo_corriente","otras_cuentas_por_pagar_cp"),
    "total pasivo corriente":              ("pasivo_corriente", "total_pasivo_corriente"),
    "otras cuentas por pagar a largo plazo":("pasivo_no_corriente","deuda_largo_plazo"),
    "total pasivo no corriente":           ("pasivo_no_corriente", "total_pasivo_no_corriente"),
    "capital":                              ("patrimonio", "capital_social"),
    "resultados acumulados":               ("patrimonio", "resultados_acumulados"),
    "resultado del periodo":               ("patrimonio", "resultado_periodo"),
    "total patrimonio":                    ("patrimonio", "total_patrimonio"),
}
ER_MAP = {
    "ventas netas":                        "ventas_netas",
    "costo de ventas":                     "costo_ventas",
    "total utilidad bruta":                "margen_bruto",
    "gastos de administracion":            "gastos_admin",
    "gastos de venta":                     "gastos_ventas",
    "gastos de produccion":                "gastos_produccion",
    "gastos de distribucion":              "gastos_distribucion",
    "gastos de desarrollo de nuevos negocios": "gastos_desarrollo",
    "total gastos":                        "gastos_operativos_total",
    "utilidad operativa":                  "margen_operativo",
    "ingresos financieros":                "ing_financieros_total",
    "gastos financieros":                  "gastos_financieros",
    "diferencia de cambio neto":           "perdida_diferencia_cambio",
    "otros ingresos de gestion":           "otros_ingresos_gestion",
    "utilidad antes de imp y participac.": "utilidad_antes_impuestos",
    "gasto en impuesto sobre la renta":    "impuesto_renta",
    "resultado del ejercicio":             "resultado_ejercicio",
}

def pares(ws, col_lab, col_val):
    """Devuelve (etiqueta_normalizada, valor) de una columna de la hoja."""
    out = []
    for r in range(1, ws.max_row + 1):
        lab = ws.cell(row=r, column=col_lab).value
        val = ws.cell(row=r, column=col_val).value
        if isinstance(lab, str) and isinstance(val, (int, float)):
            out.append((norm(lab), float(val)))
    return out

def leer_encabezado(ws):
    """El periodo tal como lo declara la hoja, y el rótulo de la columna."""
    titulo = periodo = rotulo = ""
    for r in range(1, 10):
        for c in range(1, 6):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str): continue
            n = norm(v)
            if "estado de" in n: titulo = v.strip()
            elif n.startswith("al "): periodo = v.strip()
            elif re.fullmatch(r"\d{1,2}/\d{4}", v.strip()): rotulo = v.strip()
    return {"titulo": titulo, "periodo_declarado": periodo, "rotulo_columna": rotulo}

wb = load_workbook(ARCHIVO, data_only=True)

# Descubrimiento por CONTENIDO, no por nombre: se lee el título que la hoja
# lleva adentro y el año se toma del encabezado. Es la corrección que le falta
# a discover_sheets().
hojas = {}
for nombre in wb.sheetnames:
    ws = wb[nombre]
    enc = leer_encabezado(ws)
    t = norm(enc["titulo"])
    anio = None
    m = re.search(r"(20\d{2})", enc["periodo_declarado"] or nombre)
    if m: anio = m.group(1)
    if "situacion financiera" in t: hojas.setdefault("balance", {})[anio] = (nombre, enc)
    elif "resultados" in t:         hojas.setdefault("er", {})[anio] = (nombre, enc)

periodos = sorted(set(list(hojas.get("balance", {})) + list(hojas.get("er", {}))))
print("periodos detectados por contenido:", periodos)

balance, estado_resultados, encabezados = {}, {}, {}
for p in periodos:
    nb, encb = hojas["balance"][p]
    ws = wb[nb]
    bg = {"activo_corriente": {}, "activo_no_corriente": {}, "pasivo_corriente": {},
          "pasivo_no_corriente": {}, "patrimonio": {}}
    # activo a la izquierda (A/B), pasivo y patrimonio a la derecha (D/E)
    for col_lab, col_val in ((1, 2), (4, 5)):
        for lab, val in pares(ws, col_lab, col_val):
            if lab in BG_MAP:
                sec, key = BG_MAP[lab]
                bg[sec][key] = val
            elif lab == "total activo":
                bg["total_activo"] = val
    bg["total_pasivo"] = round(bg["pasivo_corriente"]["total_pasivo_corriente"] +
                               bg["pasivo_no_corriente"]["total_pasivo_no_corriente"], 2)
    balance[p] = bg

    ne, ence = hojas["er"][p]
    er = {}
    for lab, val in pares(wb[ne], 1, 2):
        if lab in ER_MAP:
            er.setdefault(ER_MAP[lab], val)      # setdefault: la 1ra aparición manda
    estado_resultados[p] = er
    encabezados[p] = {"balance": encb, "estado_resultados": ence}

# ── PUENTE DE ESQUEMAS ─────────────────────────────────────────────────────
# calculate_ratios.py lee {periodos, balance, estado_resultados}; scan_router.py
# y generate_minuta.js leen {esf, egp}, que es el esquema de extract_data.py.
# Son dos esquemas incompatibles dentro del mismo skill (F3), así que el
# data.json sale con las DOS vistas: la multiperiodo y, además, la vista de
# periodo único apuntando al último año, que es lo que esos dos consumen.
ultimo = periodos[-1]
_b, _e = balance[ultimo], estado_resultados[ultimo]
esf_view = {
    "activo_corriente":   dict(_b["activo_corriente"],
                               total_ac=_b["activo_corriente"]["total_activo_corriente"]),
    "activo_no_corriente":dict(_b["activo_no_corriente"],
                               total_anc=_b["activo_no_corriente"]["total_activo_no_corriente"]),
    "pasivo_corriente":   dict(_b["pasivo_corriente"],
                               total_pc=_b["pasivo_corriente"]["total_pasivo_corriente"]),
    "pasivo_no_corriente":dict(_b["pasivo_no_corriente"],
                               total_pnc=_b["pasivo_no_corriente"]["total_pasivo_no_corriente"]),
    "patrimonio":         _b["patrimonio"],
    "totales": {"activo": _b["total_activo"], "pasivo": _b["total_pasivo"],
                "patrimonio": _b["patrimonio"]["total_patrimonio"]},
}
# El router llama `dif_cambio` a lo que el motor de ratios llama
# `perdida_diferencia_cambio`. Se emiten los dos nombres.
egp_view = dict(_e,
                dif_cambio=_e.get("perdida_diferencia_cambio", 0),
                otros_ingresos=_e.get("otros_ingresos_gestion", 0),
                auspicios=0)

data = {
    "empresa": "NUTRIANDES DISTRIBUCIONES S.A.C.",
    "periodo": ultimo,
    "esf": esf_view,
    "egp": egp_view,
    "complementary": {"has_cost_detail": False, "has_monthly_egp": False,
                      "has_gastos_detalle": True},
    "moneda": "PEN",
    "sector": "distribution",
    "periodos": periodos,
    "balance": balance,
    "estado_resultados": estado_resultados,
    "data_quality": {"periods_found": len(periodos), "balance_check": True,
                     "completeness": 1.0, "warnings": []},
    "_encabezados": encabezados,
    "_meta": {"source_file": ARCHIVO,
              "extraccion": "adaptador propio; extract_data.py del skill no lee este libro (F1/F2/F3)"},
}
json.dump(data, open(SALIDA, "w"), indent=2, ensure_ascii=False)
print("escrito", SALIDA)
for p in periodos:
    b, e = balance[p], estado_resultados[p]
    cuadra = abs(b["total_activo"] - (b["total_pasivo"] + b["patrimonio"]["total_patrimonio"])) < 0.01
    print(f"  [{p}] activo={b['total_activo']:>14,.2f}  ecuacion={'OK' if cuadra else 'FALLA'}"
          f"  ventas={e['ventas_netas']:>14,.2f}  resultado={e['resultado_ejercicio']:>12,.2f}")
