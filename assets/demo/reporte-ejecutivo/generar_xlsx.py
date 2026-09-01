"""
Escribe caso_demo.xlsx: el archivo de ENTRADA del demo.

Imita el layout de un juego real de EEFF peruano de una distribuidora:
balance en dos columnas (activo a la izquierda, pasivo y patrimonio a la derecha),
estado de resultados en lista vertical, encabezados con razón social y periodo.

Incluye tres defectos de calidad plantados a propósito (D1, D2, D3) para que la
minuta del skill los levante. Están marcados en el código.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from caso_sintetico import EMPRESA, A1, A2, ER, BAL

ARIAL = "Arial"
b = Font(name=ARIAL, bold=True, size=10)
n = Font(name=ARIAL, size=10)
t = Font(name=ARIAL, bold=True, size=11)
MONEY = '#,##0.00;(#,##0.00);-'


def encabezado(ws, titulo, periodo, col=1):
    for i, txt in enumerate([EMPRESA, "RUC 20XXXXXXXXX", titulo, periodo,
                             "(Expresado en soles)"], start=2):
        c = ws.cell(row=i, column=col, value=txt)
        c.font = t if i <= 3 else n
    return 8


def escribir_balance(ws, anio, rotulo_columna):
    r0 = encabezado(ws, "ESTADO DE SITUACIÓN FINANCIERA",
                    f"Al 31 de Diciembre del {anio}")
    d = BAL[anio]
    ws.cell(row=r0, column=1, value="ACTIVO").font = b
    ws.cell(row=r0, column=2, value=rotulo_columna).font = b   # <- D3 en el año 1
    ws.cell(row=r0, column=4, value="PASIVO Y PATRIMONIO").font = b
    ws.cell(row=r0, column=5, value=rotulo_columna).font = b

    r = r0 + 1
    ws.cell(row=r, column=1, value="ACTIVO CORRIENTE").font = b
    r += 1
    for k, v in d["activo_corriente"].items():
        ws.cell(row=r, column=1, value=k).font = n
        c = ws.cell(row=r, column=2, value=v); c.font = n; c.number_format = MONEY
        r += 1
    ws.cell(row=r, column=1, value="Total ACTIVO CORRIENTE").font = b
    c = ws.cell(row=r, column=2, value=d["total_ac"]); c.font = b; c.number_format = MONEY
    r += 2
    ws.cell(row=r, column=1, value="ACTIVO NO CORRIENTE").font = b
    r += 1
    for k, v in d["activo_no_corriente"].items():
        ws.cell(row=r, column=1, value=k).font = n
        c = ws.cell(row=r, column=2, value=v); c.font = n; c.number_format = MONEY
        r += 1
    ws.cell(row=r, column=1, value="Total ACTIVO NO CORRIENTE").font = b
    c = ws.cell(row=r, column=2, value=d["total_anc"]); c.font = b; c.number_format = MONEY
    r += 2
    ws.cell(row=r, column=1, value="Total ACTIVO").font = b
    c = ws.cell(row=r, column=2, value=d["total_activo"]); c.font = b; c.number_format = MONEY

    r = r0 + 1
    ws.cell(row=r, column=4, value="PASIVO CORRIENTE").font = b
    r += 1
    for k, v in d["pasivo_corriente"].items():
        ws.cell(row=r, column=4, value=k).font = n
        c = ws.cell(row=r, column=5, value=v); c.font = n; c.number_format = MONEY
        r += 1
    ws.cell(row=r, column=4, value="Total PASIVO CORRIENTE").font = b
    c = ws.cell(row=r, column=5, value=d["total_pc"]); c.font = b; c.number_format = MONEY
    r += 2
    ws.cell(row=r, column=4, value="PASIVO NO CORRIENTE").font = b
    r += 1
    for k, v in d["pasivo_no_corriente"].items():
        ws.cell(row=r, column=4, value=k).font = n
        c = ws.cell(row=r, column=5, value=v); c.font = n; c.number_format = MONEY
        r += 1
    ws.cell(row=r, column=4, value="Total PASIVO NO CORRIENTE").font = b
    c = ws.cell(row=r, column=5, value=d["total_pnc"]); c.font = b; c.number_format = MONEY
    r += 2
    ws.cell(row=r, column=4, value="PATRIMONIO").font = b
    r += 1
    for k, v in d["patrimonio"].items():
        ws.cell(row=r, column=4, value=k).font = n
        c = ws.cell(row=r, column=5, value=v); c.font = n; c.number_format = MONEY
        r += 1
    ws.cell(row=r, column=4, value="Total PATRIMONIO").font = b
    c = ws.cell(row=r, column=5, value=d["total_patrimonio"]); c.font = b; c.number_format = MONEY
    r += 2
    ws.cell(row=r, column=4, value="Total PASIVO Y PATRIMONIO").font = b
    c = ws.cell(row=r, column=5, value=d["total_activo"]); c.font = b; c.number_format = MONEY

    ws.column_dimensions['A'].width = 44
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['D'].width = 44
    ws.column_dimensions['E'].width = 18


def escribir_er(ws, anio, periodo_encabezado):
    r = encabezado(ws, "ESTADO DE RESULTADOS INTEGRALES", periodo_encabezado)
    e = ER[anio]
    ws.cell(row=r, column=1, value="Nombre de la cuenta").font = b
    ws.cell(row=r, column=2, value="Balance").font = b
    r += 2

    def linea(label, val, bold=False, fmt=True):
        nonlocal r
        c1 = ws.cell(row=r, column=1, value=label); c1.font = b if bold else n
        if val is not None:
            c2 = ws.cell(row=r, column=2, value=val)
            c2.font = b if bold else n
            if fmt: c2.number_format = MONEY
        r += 1

    linea("UTILIDAD BRUTA", e["utilidad_bruta"], bold=True)
    linea("Ventas Netas", e["ventas"])
    linea("Costo de Ventas", e["costo_ventas"])
    linea("Total UTILIDAD BRUTA", e["utilidad_bruta"], bold=True)
    r += 1
    linea("GASTOS", e["gastos_total"], bold=True)
    for k, v in e["gastos"].items():          # <- D2: en el año 2 aparece una categoría nueva
        linea(k, v)
    linea("Total GASTOS", e["gastos_total"], bold=True)
    r += 1
    linea("UTILIDAD OPERATIVA", e["utilidad_operativa"], bold=True)
    r += 1
    linea("INGRESOS FINANCIEROS", e["ingresos_financieros"])
    linea("GASTOS FINANCIEROS", e["gastos_financieros"])
    r += 1
    otros = e["diferencia_cambio"] + e["otros_ingresos_gestion"]
    linea("OTROS INGRESOS Y GASTOS", otros, bold=True)
    linea("DIFERENCIA DE CAMBIO NETO", e["diferencia_cambio"])
    linea("OTROS INGRESOS DE GESTIÓN", e["otros_ingresos_gestion"])
    linea("Total OTROS INGRESOS Y GASTOS", otros, bold=True)
    r += 1
    linea("UTILIDAD ANTES DE IMP Y PARTICIPAC.", e["uai"], bold=True)
    r += 1
    linea("Gasto en impuesto sobre la renta", e["impuesto_renta"])
    r += 1
    linea("RESULTADO DEL EJERCICIO", e["resultado"], bold=True)

    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 18


wb = Workbook()
wb.remove(wb.active)

# Año 1. D3: el título dice diciembre, el rótulo de columna dice 11/A1.
escribir_balance(wb.create_sheet(f"BG {A1}"), A1, f"11/{A1}")
escribir_er(wb.create_sheet(f"ERF {A1}"), A1, f"Al 31 de Diciembre del {A1}")

# Año 2. D1: encabezado del ER desfasado, dice octubre y los datos son de cierre.
escribir_balance(wb.create_sheet(f"BG {A2}"), A2, f"12/{A2}")
escribir_er(wb.create_sheet(f"ERF {A2}"), A2, f"Al 31/10/{A2}")

wb.save("/mnt/user-data/outputs/caso_demo.xlsx")
print("caso_demo.xlsx escrito")
